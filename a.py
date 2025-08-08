import streamlit as st
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io
import os
import tempfile
import traceback
from typing import Dict, List, Tuple, Optional

# Configure Streamlit page
st.set_page_config(
    page_title="📊 Unified Report Generator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Hide Streamlit UI components
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Custom styling */
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        text-align: center;
    }
    
    .metric-container {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #007acc;
        margin: 1rem 0;
    }
    
    .error-container {
        background: #fff5f5;
        border: 1px solid #fed7d7;
        color: #c53030;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    
    .success-container {
        background: #f0fff4;
        border: 1px solid #9ae6b4;
        color: #276749;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

###############################################################################
# -------------------------  CONFIGURATION  -----------------------------------
###############################################################################

class Config:
    """Configuration constants for the application"""
    
    # Columns to keep for slippage analysis
    KEEP_SLIPPAGE = [
        'Branch Name', 'Main Code', 'Ac Type Desc', 'Name',
        'Limit', 'Balance', 'Provision'
    ]
    
    # Provision mapping and categories - FIXED to include Watchlist
    PROV_MAP = {'G': 1, 'W': 2, 'S': 3, 'D': 4, 'B': 5}
    CAT_NAMES = {'G': 'Good', 'W': 'Watchlist', 'S': 'Substandard', 'D': 'Doubtful', 'B': 'Bad'}
    CAT_ORDER = ['Good', 'Watchlist', 'Substandard', 'Doubtful', 'Bad']
    
    # Staff loan types to exclude
    STAFF_LOANS = {
        'STAFF SOCIAL LOAN', 'STAFF VEHICLE LOAN', 'STAFF HOME LOAN',
        'STAFF FLEXIBLE LOAN', 'STAFF HOME LOAN(COF)',
        'STAFF VEHICLE FACILITY LOAN (EVF)'
    }

###############################################################################
# -------------------------  AUTHENTICATION  ----------------------------------
###############################################################################

class AuthManager:
    """Handle user authentication"""
    
    @staticmethod
    def show_login():
        """Display login form"""
        st.markdown("""
            <div class="main-header">
                <h1>🔐 Authentication Required</h1>
                <p>Please enter your credentials to access the Report Generator</p>
            </div>
        """, unsafe_allow_html=True)
        
        with st.container():
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                with st.form("login_form", clear_on_submit=False):
                    username = st.text_input("👤 Username", placeholder="Enter username")
                    password = st.text_input("🔒 Password", type="password", placeholder="Enter password")
                    
                    col_a, col_b, col_c = st.columns([1, 1, 1])
                    with col_b:
                        submitted = st.form_submit_button("🚀 Login", use_container_width=True)
                    
                    if submitted:
                        if AuthManager.validate_credentials(username, password):
                            st.session_state["authenticated"] = True
                            st.session_state["username"] = username
                            st.success("✅ Login successful!")
                            st.rerun()
                        else:
                            st.error("❌ Invalid username or password")
    
    @staticmethod
    def validate_credentials(username: str, password: str) -> bool:
        """Validate user credentials"""
        try:
            auth_config = st.secrets.get("auth", {})
            return username in auth_config and password == auth_config[username]
        except Exception:
            # Default credentials for demo (remove in production)
            return username == "admin" and password == "admin123"
    
    @staticmethod
    def show_logout_sidebar():
        """Display logout option in sidebar"""
        with st.sidebar:
            st.markdown("---")
            username = st.session_state.get("username", "User")
            st.markdown(f"👋 Welcome, **{username}**")
            
            if st.button("🚪 Logout", use_container_width=True):
                for key in ["authenticated", "username"]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()

###############################################################################
# -------------------------  DATA PROCESSING  --------------------------------
###############################################################################

class DataProcessor:
    """Handle data preprocessing and analysis"""
    
    @staticmethod
    def validate_columns(df: pd.DataFrame, required_cols: List[str], file_name: str) -> None:
        """Validate that DataFrame contains required columns"""
        df.columns = df.columns.str.strip()
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise ValueError(f"❌ {file_name} is missing columns: {', '.join(missing)}")
    
    @staticmethod
    def preprocess_slippage(df: pd.DataFrame) -> pd.DataFrame:
        """Preprocess data for slippage analysis"""
        DataProcessor.validate_columns(df, Config.KEEP_SLIPPAGE, "Slippage file")
        
        # Select and clean data
        df_clean = df[Config.KEEP_SLIPPAGE].copy()
        
        # Convert numeric columns
        for col in ['Limit', 'Balance']:
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
        
        # Remove invalid records
        df_clean = df_clean.dropna(subset=['Limit', 'Balance'])
        df_clean = df_clean[df_clean['Limit'] != 0]
        
        # Process provision codes
        df_clean['Prov_init'] = df_clean['Provision'].astype(str).str.upper().str[0]
        invalid_codes = df_clean[~df_clean['Prov_init'].isin(Config.PROV_MAP)]
        
        if not invalid_codes.empty:
            raise ValueError(f"❌ Invalid provision codes found: {invalid_codes['Prov_init'].unique()}")
        
        df_clean['Prov_rank'] = df_clean['Prov_init'].map(Config.PROV_MAP)
        df_clean['Prov_cat'] = df_clean['Prov_init'].map(Config.CAT_NAMES)
        
        return df_clean
    
    @staticmethod
    def preprocess_comparison(df: pd.DataFrame) -> pd.DataFrame:
        """Preprocess data for balance comparison"""
        df_clean = df.copy()
        
        # Clean account type descriptions
        df_clean['Ac Type Desc'] = df_clean['Ac Type Desc'].str.strip().str.upper()
        
        # Exclude staff loans
        staff_loans_upper = {loan.upper() for loan in Config.STAFF_LOANS}
        df_clean = df_clean[~df_clean['Ac Type Desc'].isin(staff_loans_upper)]
        
        # Remove zero limits and total rows
        df_clean = df_clean[df_clean['Limit'] != 0]
        df_clean = df_clean[~df_clean['Main Code'].isin({'AcType Total', 'Grand Total'})]
        
        return df_clean
    
    @staticmethod
    def detect_slippage(df_prev: pd.DataFrame, df_curr: pd.DataFrame) -> pd.DataFrame:
        """Detect account slippage between periods"""
        # Prepare previous period data
        prev_data = df_prev[['Main Code', 'Prov_rank', 'Prov_cat']].rename(
            columns={'Prov_rank': 'rank_prev', 'Prov_cat': 'cat_prev'}
        )
        
        # Prepare current period data
        curr_data = df_curr[['Main Code', 'Prov_rank', 'Prov_cat']].rename(
            columns={'Prov_rank': 'rank_curr', 'Prov_cat': 'cat_curr'}
        )
        
        # Merge data on Main Code
        merged = pd.merge(prev_data, curr_data, on='Main Code', how='inner')
        
        # Get full current period data for matched accounts
        full_data = (
            df_curr[df_curr['Main Code'].isin(merged['Main Code'])]
            .merge(merged[['Main Code', 'rank_prev', 'cat_prev']], on='Main Code')
        )
        
        # Determine movement type
        def get_movement(row):
            if row['Prov_rank'] > row['rank_prev']:
                return "Slippage"
            elif row['Prov_rank'] < row['rank_prev']:
                return "Upgrade"
            else:
                return "Stable"
        
        full_data['Movement'] = full_data.apply(get_movement, axis=1)
        
        # Select final columns - FIXED to include previous category
        result_cols = [
            'Branch Name', 'Main Code', 'Ac Type Desc', 'Name',
            'Limit', 'Balance', 'cat_prev', 'Prov_cat', 'Movement'
        ]
        
        return full_data[result_cols].rename(columns={'Prov_cat': 'cat_curr'})

###############################################################################
# -------------------------  REPORT GENERATION  ------------------------------
###############################################################################

class ReportGenerator:
    """Generate Excel reports with various analyses"""
    
    @staticmethod
    def create_category_matrix(df: pd.DataFrame, group_col: Optional[str] = None) -> pd.DataFrame:
        """Create category transition matrix - FIXED to show counts by previous category"""
        if group_col:
            # Create pivot table showing current category counts grouped by previous category and group column
            matrix = (
                df.pivot_table(
                    index=group_col,
                    columns='cat_prev',  # Group by PREVIOUS category
                    values='Main Code',  # Count accounts
                    aggfunc='count',
                    fill_value=0
                )
                .reindex(columns=Config.CAT_ORDER, fill_value=0)  # Include all categories including Watchlist
                .astype(int)
                .reset_index()
            )
        else:
            # Overall summary
            matrix = (
                df.groupby('cat_prev')['Main Code']
                .count()
                .reindex(Config.CAT_ORDER, fill_value=0)
                .to_frame('Count')
                .T
                .reset_index(drop=True)
            )
        
        return matrix
    
    @staticmethod
    def create_movement_summary(df: pd.DataFrame) -> pd.DataFrame:
        """Create movement summary by type"""
        movement_summary = (
            df.groupby(['Movement', 'cat_prev', 'cat_curr'])
            .agg({
                'Main Code': 'count',
                'Balance': 'sum',
                'Limit': 'sum'
            })
            .round(2)
            .reset_index()
            .rename(columns={
                'Main Code': 'Count',
                'Balance': 'Total_Balance',
                'Limit': 'Total_Limit'
            })
        )
        return movement_summary
    
    @staticmethod
    def create_balance_comparison(df_prev: pd.DataFrame, df_curr: pd.DataFrame, writer) -> Dict:
        """Create balance comparison analysis"""
        # Validate required columns
        required_cols = {'Main Code', 'Balance'}
        for df, name in [(df_prev, 'Previous'), (df_curr, 'Current')]:
            missing = required_cols - set(df.columns)
            if missing:
                raise ValueError(f"❌ {name} file missing columns: {', '.join(missing)}")
        
        # Get account codes from each period
        prev_codes = set(df_prev['Main Code'])
        curr_codes = set(df_curr['Main Code'])
        
        # Identify settled and new accounts
        settled_accounts = df_prev[df_prev['Main Code'].isin(prev_codes - curr_codes)]
        new_accounts = df_curr[df_curr['Main Code'].isin(curr_codes - prev_codes)]
        
        # Analyze continuing accounts
        continuing_accounts = pd.merge(
            df_prev[['Main Code', 'Balance']].rename(columns={'Balance': 'Prev_Bal'}),
            df_curr[['Main Code', 'Branch Name', 'Name', 'Ac Type Desc', 'Balance']]
            .rename(columns={'Balance': 'Curr_Bal'}),
            on='Main Code'
        )
        continuing_accounts['Change'] = continuing_accounts['Curr_Bal'] - continuing_accounts['Prev_Bal']
        
        # Create reconciliation summary
        reconciliation = pd.DataFrame({
            'Description': ['Opening Balance', 'Settled Accounts', 'New Accounts', 'Balance Changes', 'Adjusted Balance', 'Closing Balance'],
            'Amount': [
                df_prev['Balance'].sum(),
                -settled_accounts['Balance'].sum(),
                new_accounts['Balance'].sum(),
                continuing_accounts['Change'].sum(),
                0,  # Will be calculated
                df_curr['Balance'].sum()
            ],
            'Account Count': [
                len(prev_codes),
                -len(prev_codes - curr_codes),
                len(curr_codes - prev_codes),
                "",
                "",
                len(curr_codes)
            ]
        })
        
        # Calculate adjusted balance
        reconciliation.at[4, 'Amount'] = reconciliation.loc[0:4, 'Amount'].sum()
        
        # Write to Excel sheets
        settled_accounts.to_excel(writer, sheet_name='Settled_Accounts', index=False)
        new_accounts.to_excel(writer, sheet_name='New_Accounts', index=False)
        continuing_accounts[['Main Code', 'Ac Type Desc', 'Branch Name', 'Name', 'Curr_Bal', 'Prev_Bal', 'Change']].to_excel(
            writer, sheet_name='Account_Movement', index=False
        )
        reconciliation.to_excel(writer, sheet_name='Balance_Reconciliation', index=False)
        
        # Return summary statistics
        return {
            'total_prev': df_prev['Balance'].sum(),
            'total_curr': df_curr['Balance'].sum(),
            'net_change': df_curr['Balance'].sum() - df_prev['Balance'].sum(),
            'settled_count': len(settled_accounts),
            'new_count': len(new_accounts),
            'continuing_count': len(continuing_accounts)
        }
    
    @staticmethod
    def create_pivot_comparison(df_prev: pd.DataFrame, df_curr: pd.DataFrame, 
                               group_by: str, writer, sheet_name: str) -> None:
        """Create pivot comparison analysis"""
        # Group previous period data
        prev_grouped = (
            df_prev.groupby(group_by)
            .agg(Prev_Sum=('Balance', 'sum'), Prev_Count=(group_by, 'count'))
        )
        
        # Group current period data
        curr_grouped = (
            df_curr.groupby(group_by)
            .agg(Curr_Sum=('Balance', 'sum'), Curr_Count=(group_by, 'count'))
        )
        
        # Merge and calculate changes
        comparison = prev_grouped.join(curr_grouped, how='outer').fillna(0)
        comparison['Balance_Change'] = comparison['Curr_Sum'] - comparison['Prev_Sum']
        comparison['Count_Change'] = comparison['Curr_Count'] - comparison['Prev_Count']
        comparison['Pct_Change'] = (
            (comparison['Balance_Change'] / comparison['Prev_Sum'].replace(0, pd.NA) * 100)
            .fillna(0)
            .round(2)
        )
        
        # Add total row
        totals = comparison.sum(numeric_only=True)
        totals.name = 'TOTAL'
        totals['Pct_Change'] = (
            (totals['Balance_Change'] / totals['Prev_Sum'] * 100) 
            if totals['Prev_Sum'] else 0
        )
        
        final_comparison = pd.concat([comparison, totals.to_frame().T]).reset_index()
        
        # Write to Excel with formatting
        final_comparison.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the total row in bold
        ws = writer.sheets[sheet_name]
        last_row = len(final_comparison) + 1
        for col in range(1, len(final_comparison.columns) + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True)

###############################################################################
# -------------------------  EXCEL UTILITIES  --------------------------------
###############################################################################

class ExcelFormatter:
    """Handle Excel formatting and styling"""
    
    @staticmethod
    def autofit_columns(writer):
        """Auto-fit column widths in all sheets"""
        for sheet_name, worksheet in writer.sheets.items():
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def format_headers(writer):
        """Apply formatting to header rows"""
        header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
        header_font = Font(bold=True)
        
        for worksheet in writer.sheets.values():
            for cell in worksheet[1]:  # First row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

###############################################################################
# -------------------------  MAIN APPLICATION  -------------------------------
###############################################################################

def create_summary_statistics(slippage_df: pd.DataFrame) -> Dict:
    """Create summary statistics for display"""
    total_accounts = len(slippage_df)
    
    movement_counts = slippage_df['Movement'].value_counts()
    
    return {
        'total_accounts': total_accounts,
        'slippage_count': movement_counts.get('Slippage', 0),
        'upgrade_count': movement_counts.get('Upgrade', 0),
        'stable_count': movement_counts.get('Stable', 0),
        'slippage_rate': (movement_counts.get('Slippage', 0) / total_accounts * 100) if total_accounts > 0 else 0
    }

def main_app():
    """Main application interface"""
    
    # Show logout option
    AuthManager.show_logout_sidebar()
    
    # Main header
    st.markdown("""
        <div class="main-header">
            <h1>📊 Unified Report Generator</h1>
            <p>Upload Previous and Current Excel files to generate comprehensive analysis reports</p>
        </div>
    """, unsafe_allow_html=True)
    
    # File upload section
    st.subheader("📁 File Upload")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📅 Previous Period File**")
        prev_file = st.file_uploader(
            "Choose previous period Excel file",
            type=['xlsx', 'xls'],
            key="prev_file"
        )
    
    with col2:
        st.markdown("**📅 Current Period File**")
        curr_file = st.file_uploader(
            "Choose current period Excel file", 
            type=['xlsx', 'xls'],
            key="curr_file"
        )
    
    # Process files if both uploaded
    if prev_file and curr_file:
        
        # Analysis options
        st.subheader("⚙️ Analysis Options")
        
        analysis_options = st.multiselect(
            "Select analysis types to include:",
            options=[
                "Slippage Analysis",
                "Balance Comparison", 
                "Branch Comparison",
                "Account Type Comparison"
            ],
            default=[
                "Slippage Analysis",
                "Balance Comparison",
                "Branch Comparison", 
                "Account Type Comparison"
            ]
        )
        
        # Generate report button
        if st.button("🚀 Generate Comprehensive Report", type="primary", use_container_width=True):
            
            with st.spinner("🔄 Processing files and generating report..."):
                try:
                    # Save uploaded files temporarily
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_prev, \
                         tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_curr:
                        
                        tmp_prev.write(prev_file.getbuffer())
                        tmp_curr.write(curr_file.getbuffer())
                        tmp_prev_path = tmp_prev.name
                        tmp_curr_path = tmp_curr.name
                    
                    # Read Excel files
                    df_prev_raw = pd.read_excel(tmp_prev_path)
                    df_curr_raw = pd.read_excel(tmp_curr_path)
                    
                    # Create output buffer
                    output_buffer = io.BytesIO()
                    
                    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                        
                        # Slippage Analysis
                        if "Slippage Analysis" in analysis_options:
                            df_prev_slippage = DataProcessor.preprocess_slippage(df_prev_raw)
                            df_curr_slippage = DataProcessor.preprocess_slippage(df_curr_raw)
                            
                            slippage_data = DataProcessor.detect_slippage(df_prev_slippage, df_curr_slippage)
                            
                            # Write slippage data
                            slippage_data.to_excel(writer, sheet_name='Slippage_Detail', index=False)
                            
                            # Create summaries - FIXED to show previous categories properly
                            branch_matrix = ReportGenerator.create_category_matrix(slippage_data, 'Branch Name')
                            actype_matrix = ReportGenerator.create_category_matrix(slippage_data, 'Ac Type Desc')
                            movement_summary = ReportGenerator.create_movement_summary(slippage_data)
                            
                            # Write summaries
                            branch_matrix.to_excel(writer, sheet_name='Summary_Branch', index=False)
                            actype_matrix.to_excel(writer, sheet_name='Summary_AcType', index=False)
                            movement_summary.to_excel(writer, sheet_name='Movement_Summary', index=False)
                        
                        # Balance and other comparisons
                        if any(opt in analysis_options for opt in ["Balance Comparison", "Branch Comparison", "Account Type Comparison"]):
                            df_prev_comp = DataProcessor.preprocess_comparison(df_prev_raw)
                            df_curr_comp = DataProcessor.preprocess_comparison(df_curr_raw)
                            
                            if "Balance Comparison" in analysis_options:
                                balance_stats = ReportGenerator.create_balance_comparison(df_prev_comp, df_curr_comp, writer)
                            
                            if "Account Type Comparison" in analysis_options:
                                ReportGenerator.create_pivot_comparison(
                                    df_prev_comp, df_curr_comp, 'Ac Type Desc', writer, 'AcType_Comparison'
                                )
                            
                            if "Branch Comparison" in analysis_options:
                                ReportGenerator.create_pivot_comparison(
                                    df_prev_comp, df_curr_comp, 'Branch Name', writer, 'Branch_Comparison'
                                )
                        
                        # Apply formatting
                        ExcelFormatter.autofit_columns(writer)
                        ExcelFormatter.format_headers(writer)
                    
                    output_buffer.seek(0)
                    
                    # Show success message and summary stats
                    st.success("✅ Report generated successfully!")
                    
                    # Display summary statistics if slippage analysis was performed
                    if "Slippage Analysis" in analysis_options:
                        stats = create_summary_statistics(slippage_data)
                        
                        st.subheader("📈 Summary Statistics")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Total Accounts", f"{stats['total_accounts']:,}")
                        with col2:
                            st.metric("Slipped Accounts", f"{stats['slippage_count']:,}")
                        with col3:
                            st.metric("Upgraded Accounts", f"{stats['upgrade_count']:,}")
                        with col4:
                            st.metric("Slippage Rate", f"{stats['slippage_rate']:.1f}%")
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Comprehensive Report",
                        data=output_buffer,
                        file_name=f"unified_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    # Clean up temporary files
                    try:
                        os.unlink(tmp_prev_path)
                        os.unlink(tmp_curr_path)
                    except:
                        pass
                        
                except Exception as e:
                    st.error("❌ An error occurred while processing the files")
                    
                    with st.expander("🔍 View Error Details"):
                        st.code(f"Error: {str(e)}")
                        st.code(traceback.format_exc())

def main():
    """Main entry point"""
    # Initialize authentication state
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
    
    # Show login or main app
    if not st.session_state["authenticated"]:
        AuthManager.show_login()
    else:
        main_app()

if __name__ == "__main__":
    main()
